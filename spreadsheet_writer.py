from pathlib import Path
from openpyxl import Workbook, load_workbook


def write_result_to_spreadsheet(marking_result, output_folder):
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)

    spreadsheet_path = output_folder / "class_results.xlsx"

    student_name = marking_result.get("student_name", "Student")
    overall_grade = marking_result.get("overall_grade", "")
    criteria_results = marking_result.get("criteria_results", [])

    if spreadsheet_path.exists():
        workbook = load_workbook(spreadsheet_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        headers = ["Student Name"]
        sheet.append(headers)

    for criterion in criteria_results:
        criterion_name = criterion.get("criterion", "Criterion")
        grade_header = f"{criterion_name} Grade"

        if grade_header not in headers:
            headers.append(grade_header)
            sheet.cell(row=1, column=len(headers)).value = grade_header

    if "Overall Grade" not in headers:
        headers.append("Overall Grade")
        sheet.cell(row=1, column=len(headers)).value = "Overall Grade"

    row_data = [""] * len(headers)
    row_data[headers.index("Student Name")] = student_name

    for criterion in criteria_results:
        criterion_name = criterion.get("criterion", "Criterion")
        grade = criterion.get("grade", "")
        grade_header = f"{criterion_name} Grade"

        row_data[headers.index(grade_header)] = grade

    row_data[headers.index("Overall Grade")] = overall_grade

    sheet.append(row_data)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 30)

    workbook.save(spreadsheet_path)

    return spreadsheet_path
